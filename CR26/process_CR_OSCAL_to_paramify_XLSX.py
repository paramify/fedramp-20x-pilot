#!/usr/bin/env python3
"""
Convert FedRAMP Consolidated Rules 2026 (CR26) OSCAL catalog and profiles to
Paramify custom catalog / custom profile spreadsheet (XLSX) format.

Column layout (Paramify custom catalog / profile style; soc2_to_nist_aggregated.xlsx
was used only as a structural reference for family/subfamily columns):
  - Control ID
  - Name
  - Description
  - Control Family
  - Control Subfamily

Outputs (default, beside this script):
  - FedRAMP_CR26_custom_catalog.xlsx   — all catalog controls
  - FedRAMP_20x_custom_profile.xlsx    — controls in FedRAMP_20x_profile.json
  - FedRAMP_rev5_custom_profile.xlsx   — controls in FedRAMP_rev5_profile.json

Requires: openpyxl  (pip install openpyxl)
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any, Iterator

try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    openpyxl = None  # type: ignore
    Workbook = None  # type: ignore

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_CATALOG = SCRIPT_DIR / "FedRAMP_CR26_catalog.json"
DEFAULT_PROFILE_20X = SCRIPT_DIR / "FedRAMP_20x_profile.json"
DEFAULT_PROFILE_REV5 = SCRIPT_DIR / "FedRAMP_rev5_profile.json"

DEFAULT_CATALOG_OUT = SCRIPT_DIR / "FedRAMP_CR26_custom_catalog.xlsx"
DEFAULT_PROFILE_20X_OUT = SCRIPT_DIR / "FedRAMP_20x_custom_profile.xlsx"
DEFAULT_PROFILE_REV5_OUT = SCRIPT_DIR / "FedRAMP_rev5_custom_profile.xlsx"

PARAMIFY_COLUMNS = [
    "Control ID",
    "Name",
    "Description",
    "Control Family",
    "Control Subfamily",
]


def clean_prose(text: str) -> str:
    """Normalize OSCAL prose for Paramify import (aligned with process_frmr_to_oscal)."""
    if not text:
        return ""
    text = text.strip()
    if text.startswith('"') and text.endswith('"'):
        text = text[1:-1]
    if text.startswith("'") and text.endswith("'"):
        text = text[1:-1]
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)
    text = re.sub(r"_([A-Za-z0-9][A-Za-z0-9\s]*[A-Za-z0-9])_", r"\1", text)
    text = re.sub(r"\b_([A-Za-z0-9][A-Za-z0-9\s]*[A-Za-z0-9])\b", r"\1", text)
    text = re.sub(r"\b([A-Za-z0-9][A-Za-z0-9\s]*[A-Za-z0-9])_\b", r"\1", text)
    return text


def paramify_headers() -> list[str]:
    return list(PARAMIFY_COLUMNS)


def build_description(control: dict[str, Any]) -> str:
    """Statement first, then guidance / notes."""
    parts = control.get("parts") or []
    statement = ""
    extras: list[str] = []
    for part in parts:
        prose = clean_prose(part.get("prose") or "")
        if not prose:
            continue
        name = (part.get("name") or "").lower()
        if name == "statement" and not statement:
            statement = prose
        else:
            extras.append(prose)
    if statement and extras:
        return statement + "\n\n" + "\n\n".join(extras)
    return statement or "\n\n".join(extras)


def iter_catalog_controls(catalog_doc: dict[str, Any]) -> Iterator[dict[str, Any]]:
    """Walk OSCAL catalog groups and yield control records with family metadata."""

    def walk(
        groups: list[dict[str, Any]] | None,
        path_titles: list[str],
        path_ids: list[str],
    ) -> Iterator[dict[str, Any]]:
        for group in groups or []:
            gid = group.get("id") or ""
            title = group.get("title") or gid
            new_titles = path_titles + [title]
            new_ids = path_ids + [gid]
            for control in group.get("controls") or []:
                family = ""
                subfamily = ""
                if len(new_titles) >= 3:
                    family = new_titles[1]
                    subfamily = new_titles[2]
                elif len(new_titles) == 2:
                    family = new_titles[0]
                    subfamily = new_titles[1]
                yield {
                    "control": control,
                    "control_id": control.get("id") or "",
                    "title": clean_prose(control.get("title") or "") or control.get("id", ""),
                    "description": build_description(control),
                    "family": family,
                    "subfamily": subfamily,
                    "class": control.get("class") or "",
                    "path_ids": new_ids,
                }
            yield from walk(group.get("groups"), new_titles, new_ids)

    catalog = catalog_doc.get("catalog") or catalog_doc
    yield from walk(catalog.get("groups"), [], [])


def load_catalog_index(catalog_path: Path) -> dict[str, dict[str, Any]]:
    with open(catalog_path, encoding="utf-8") as f:
        doc = json.load(f)
    index: dict[str, dict[str, Any]] = {}
    for row in iter_catalog_controls(doc):
        cid = row["control_id"]
        if cid:
            index[cid] = row
    return index


def profile_control_ids(profile_path: Path) -> list[str]:
    with open(profile_path, encoding="utf-8") as f:
        doc = json.load(f)
    profile = doc.get("profile") or doc
    ids: list[str] = []
    for imp in profile.get("imports") or []:
        for block in imp.get("include-controls") or []:
            for cid in block.get("with-ids") or []:
                if cid:
                    ids.append(cid)
    return ids


def row_to_cells(record: dict[str, Any]) -> list[str]:
    """Map internal record to Paramify column order."""
    return [
        record["control_id"],
        record["title"],
        record["description"],
        record["family"],
        record["subfamily"],
    ]


def write_paramify_xlsx(
    output_path: Path,
    headers: list[str],
    records: list[dict[str, Any]],
    sheet_name: str,
) -> None:
    if openpyxl is None or Workbook is None:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(headers)
    for record in records:
        ws.append(row_to_cells(record))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_catalog_rows(catalog_path: Path) -> list[dict[str, Any]]:
    with open(catalog_path, encoding="utf-8") as f:
        doc = json.load(f)
    rows = list(iter_catalog_controls(doc))
    rows.sort(key=lambda r: (r["path_ids"], r["control_id"]))
    return rows


def build_profile_rows(catalog_index: dict[str, dict[str, Any]], profile_path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    missing: list[str] = []
    for cid in profile_control_ids(profile_path):
        rec = catalog_index.get(cid)
        if not rec:
            missing.append(cid)
            continue
        rows.append(rec)
    if missing:
        raise ValueError(
            f"Profile {profile_path.name} references {len(missing)} control(s) "
            f"not found in catalog: {missing[:5]}{'...' if len(missing) > 5 else ''}"
        )
    return rows


def validate_xlsx(
    output_path: Path,
    expected_headers: list[str],
    expected_row_count: int,
    expected_control_ids: set[str] | None = None,
) -> list[str]:
    """Return list of validation errors (empty if OK)."""
    errors: list[str] = []
    if openpyxl is None:
        errors.append("openpyxl not installed — cannot validate XLSX")
        return errors
    if not output_path.is_file():
        errors.append(f"Missing output file: {output_path}")
        return errors

    wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(max_row=1, values_only=True))
    headers = [str(c).strip() if c is not None else "" for c in header_row]
    if headers != expected_headers:
        errors.append(
            f"{output_path.name}: header mismatch.\n"
            f"  expected: {expected_headers}\n"
            f"  got:      {headers}"
        )

    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    if len(data_rows) != expected_row_count:
        errors.append(
            f"{output_path.name}: expected {expected_row_count} data rows, got {len(data_rows)}"
        )

    ids_in_sheet: set[str] = set()
    for idx, row in enumerate(data_rows, start=2):
        if not row or row[0] is None:
            errors.append(f"{output_path.name}: row {idx} missing control ID")
            continue
        cid = str(row[0]).strip()
        ids_in_sheet.add(cid)
        if len(row) < 3 or not (row[1] and str(row[1]).strip()):
            errors.append(f"{output_path.name}: row {idx} ({cid}) missing Name")
        if len(row) < 3 or not (row[2] and str(row[2]).strip()):
            errors.append(f"{output_path.name}: row {idx} ({cid}) missing Description")

    if expected_control_ids is not None and ids_in_sheet != expected_control_ids:
        extra = ids_in_sheet - expected_control_ids
        missing = expected_control_ids - ids_in_sheet
        if missing:
            errors.append(f"{output_path.name}: missing IDs in sheet: {sorted(missing)[:5]}...")
        if extra:
            errors.append(f"{output_path.name}: unexpected IDs in sheet: {sorted(extra)[:5]}...")

    wb.close()
    return errors


def run_conversion(
    catalog_path: Path,
    profile_20x_path: Path,
    profile_rev5_path: Path,
    catalog_out: Path,
    profile_20x_out: Path,
    profile_rev5_out: Path,
) -> dict[str, Any]:
    headers = paramify_headers()

    catalog_rows = build_catalog_rows(catalog_path)
    catalog_index = {r["control_id"]: r for r in catalog_rows}
    profile_20x_rows = build_profile_rows(catalog_index, profile_20x_path)
    profile_rev5_rows = build_profile_rows(catalog_index, profile_rev5_path)

    write_paramify_xlsx(
        catalog_out,
        headers,
        catalog_rows,
        sheet_name="FedRAMP_CR26_catalog",
    )
    write_paramify_xlsx(
        profile_20x_out,
        headers,
        profile_20x_rows,
        sheet_name="FedRAMP_20x_profile",
    )
    write_paramify_xlsx(
        profile_rev5_out,
        headers,
        profile_rev5_rows,
        sheet_name="FedRAMP_rev5_profile",
    )

    ids_20x = set(profile_control_ids(profile_20x_path))
    ids_rev5 = set(profile_control_ids(profile_rev5_path))
    all_catalog_ids = {r["control_id"] for r in catalog_rows}

    validation_errors: list[str] = []
    validation_errors.extend(
        validate_xlsx(catalog_out, headers, len(catalog_rows), all_catalog_ids)
    )
    validation_errors.extend(
        validate_xlsx(profile_20x_out, headers, len(profile_20x_rows), ids_20x)
    )
    validation_errors.extend(
        validate_xlsx(profile_rev5_out, headers, len(profile_rev5_rows), ids_rev5)
    )

    return {
        "headers": headers,
        "catalog_rows": len(catalog_rows),
        "profile_20x_rows": len(profile_20x_rows),
        "profile_rev5_rows": len(profile_rev5_rows),
        "catalog_out": catalog_out,
        "profile_20x_out": profile_20x_out,
        "profile_rev5_out": profile_rev5_out,
        "validation_errors": validation_errors,
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Convert CR26 OSCAL catalog/profiles to Paramify custom catalog/profile XLSX."
    )
    parser.add_argument("--catalog", type=Path, default=DEFAULT_CATALOG, help="OSCAL catalog JSON")
    parser.add_argument("--profile-20x", type=Path, default=DEFAULT_PROFILE_20X, help="20x OSCAL profile JSON")
    parser.add_argument("--profile-rev5", type=Path, default=DEFAULT_PROFILE_REV5, help="Rev5 OSCAL profile JSON")
    parser.add_argument("--catalog-out", type=Path, default=DEFAULT_CATALOG_OUT)
    parser.add_argument("--profile-20x-out", type=Path, default=DEFAULT_PROFILE_20X_OUT)
    parser.add_argument("--profile-rev5-out", type=Path, default=DEFAULT_PROFILE_REV5_OUT)
    parser.add_argument(
        "--validate-only",
        action="store_true",
        help="Validate existing output XLSX files without regenerating",
    )
    args = parser.parse_args()

    if openpyxl is None:
        print("ERROR: openpyxl is not installed. Run: pip install openpyxl", file=sys.stderr)
        return 1

    for p in (args.catalog, args.profile_20x, args.profile_rev5):
        if not p.is_file():
            print(f"ERROR: missing input file: {p}", file=sys.stderr)
            return 1

    print("=" * 60)
    print("CR26 OSCAL → Paramify XLSX")
    print("=" * 60)

    if args.validate_only:
        headers = paramify_headers()
        catalog_index = load_catalog_index(args.catalog)
        catalog_ids = set(catalog_index.keys())
        results = {
            "validation_errors": [],
            "headers": headers,
        }
        results["validation_errors"].extend(
            validate_xlsx(args.catalog_out, headers, len(catalog_ids), catalog_ids)
        )
        results["validation_errors"].extend(
            validate_xlsx(
                args.profile_20x_out,
                headers,
                len(profile_control_ids(args.profile_20x)),
                set(profile_control_ids(args.profile_20x)),
            )
        )
        results["validation_errors"].extend(
            validate_xlsx(
                args.profile_rev5_out,
                headers,
                len(profile_control_ids(args.profile_rev5)),
                set(profile_control_ids(args.profile_rev5)),
            )
        )
    else:
        results = run_conversion(
            args.catalog,
            args.profile_20x,
            args.profile_rev5,
            args.catalog_out,
            args.profile_20x_out,
            args.profile_rev5_out,
        )
        print("\nColumns:")
        for h in results["headers"]:
            print(f"  - {h}")
        print(f"\nWrote catalog ({results['catalog_rows']} controls): {results['catalog_out']}")
        print(f"Wrote 20x profile ({results['profile_20x_rows']} controls): {results['profile_20x_out']}")
        print(f"Wrote rev5 profile ({results['profile_rev5_rows']} controls): {results['profile_rev5_out']}")

    if results["validation_errors"]:
        print("\nVALIDATION FAILED:")
        for err in results["validation_errors"]:
            print(f"  - {err}")
        return 1

    print("\nValidation passed: headers and row counts OK; control IDs match source.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
