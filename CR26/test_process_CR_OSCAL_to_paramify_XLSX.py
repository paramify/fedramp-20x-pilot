#!/usr/bin/env python3
"""Tests for process_CR_OSCAL_to_paramify_XLSX.py"""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
PROCESSOR = SCRIPT_DIR / "process_CR_OSCAL_to_paramify_XLSX.py"
PROFILE_20X = SCRIPT_DIR / "FedRAMP_20x_profile.json"
PROFILE_REV5 = SCRIPT_DIR / "FedRAMP_rev5_profile.json"

EXPECTED_HEADERS = [
    "Control ID",
    "Name",
    "Description",
    "Control Family",
    "Control Subfamily",
]


def _profile_ids(path: Path) -> set[str]:
    doc = json.loads(path.read_text(encoding="utf-8"))
    ids: set[str] = set()
    for imp in doc["profile"].get("imports") or []:
        for block in imp.get("include-controls") or []:
            ids.update(block.get("with-ids") or [])
    return ids


def test_expected_headers_constant():
    from process_CR_OSCAL_to_paramify_XLSX import PARAMIFY_COLUMNS

    assert PARAMIFY_COLUMNS == EXPECTED_HEADERS
    assert "Soc" not in PARAMIFY_COLUMNS[0]
    assert "SOC" not in PARAMIFY_COLUMNS[0].upper() or PARAMIFY_COLUMNS[0] == "Control ID"


def test_end_to_end_conversion_and_validation():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return

    result = subprocess.run(
        [sys.executable, str(PROCESSOR)],
        cwd=SCRIPT_DIR,
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0, result.stderr or result.stdout

    catalog_out = SCRIPT_DIR / "FedRAMP_CR26_custom_catalog.xlsx"
    profile_20x_out = SCRIPT_DIR / "FedRAMP_20x_custom_profile.xlsx"
    profile_rev5_out = SCRIPT_DIR / "FedRAMP_rev5_custom_profile.xlsx"
    for path in (catalog_out, profile_20x_out, profile_rev5_out):
        assert path.is_file(), f"Missing {path}"

    cat_wb = openpyxl.load_workbook(catalog_out, read_only=True, data_only=True)
    headers = list(next(cat_wb.active.iter_rows(max_row=1, values_only=True)))
    assert headers == EXPECTED_HEADERS
    cat_rows = list(cat_wb.active.iter_rows(min_row=2, values_only=True))
    cat_wb.close()
    assert len(cat_rows) == 275

    p20_wb = openpyxl.load_workbook(profile_20x_out, read_only=True, data_only=True)
    assert list(next(p20_wb.active.iter_rows(max_row=1, values_only=True))) == EXPECTED_HEADERS
    p20_ids = {str(r[0]) for r in p20_wb.active.iter_rows(min_row=2, values_only=True) if r[0]}
    p20_wb.close()
    assert p20_ids == _profile_ids(PROFILE_20X)
    assert len(p20_ids) == 269

    p5_wb = openpyxl.load_workbook(profile_rev5_out, read_only=True, data_only=True)
    assert list(next(p5_wb.active.iter_rows(max_row=1, values_only=True))) == EXPECTED_HEADERS
    p5_ids = {str(r[0]) for r in p5_wb.active.iter_rows(min_row=2, values_only=True) if r[0]}
    p5_wb.close()
    assert p5_ids == _profile_ids(PROFILE_REV5)
    assert len(p5_ids) == 223


if __name__ == "__main__":
    test_expected_headers_constant()
    test_end_to_end_conversion_and_validation()
    print("All tests passed.")
